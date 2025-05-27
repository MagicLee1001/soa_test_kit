# -*- coding: utf-8 -*-
# @Author  : Li Kun
# @Time    : 2023/10/23 11:18
# @File    : tester.py

import os
import re
import time
import json
import traceback
import requests
import math
import threading
import ast
from typing import List, Optional
from datetime import datetime
from decimal import Decimal
from json.decoder import JSONDecodeError
from openpyxl import load_workbook
from prettytable import PrettyTable
from settings import env
from connector import *
from runner.log import logger
from runner.variable import Variable
from runner.simulator import VehicleModeDiagnostic, DoIPMonitorThread
from runner.cloud import CloudConnector


class TestStep:
    """
    单条(one row)测试步骤的数据类型
    """
    def __init__(
        self,
        pre_condition: str,
        actions: list,
        wait_condition: str,
        pass_condition: list,
        hold_condition: str,
        row_number: int,
        heading: str
    ):
        self.pre_condition = pre_condition
        self.actions = actions
        self.heading = heading
        self.wait_condition = wait_condition
        self.pass_condition = pass_condition
        self.hold_condition = hold_condition
        self.row_number = row_number
        self.heading = heading
        self.evaluation_condition = []  # 实际运行的结果
        self.time_time = ''
        self.step_ret = True


class TestInfo:
    """
    一条测试用例的数据类型
    """
    def __init__(
        self,
        tc_name: str,
        tc_steps: 'List[TestStep]',
        tc_ret: bool,
        tc_title=''
    ):
        self.tc_name = tc_name
        self.tc_title = tc_title
        self.tc_steps = tc_steps
        self.tc_ret = tc_ret
        self.test_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


class ThreadSafeProperty:
    def __init__(self, value=None):
        self._value = value
        self._lock = threading.Lock()

    def __get__(self, instance, owner):
        with self._lock:
            return self._value

    def __set__(self, instance, value):
        with self._lock:
            self._value = value


class TestHandle:
    # 类变量全部设置成线程安全模式
    case_seq = ThreadSafeProperty(1)
    chat_id = ThreadSafeProperty('')
    card_temp_id = ThreadSafeProperty('')
    notice_url = ThreadSafeProperty('')
    vin = ThreadSafeProperty('')
    title = ThreadSafeProperty('')
    case_name = ThreadSafeProperty('')
    report_dir = ThreadSafeProperty('')
    report_html_path = ThreadSafeProperty('')
    result_html_path = []
    template_bg = ThreadSafeProperty('')
    start_time = ThreadSafeProperty('')
    cost_time = ThreadSafeProperty('0 s')
    result_str = ThreadSafeProperty('')
    total_num = ThreadSafeProperty(0)
    pass_num = ThreadSafeProperty(0)
    fail_num = ThreadSafeProperty(0)
    error_num = ThreadSafeProperty(0)
    pass_rate = ThreadSafeProperty('0%')
    test_detail = ThreadSafeProperty('')

    # 给QT界面用
    all_case_num = ThreadSafeProperty(0)
    run_state = ThreadSafeProperty('未执行')
    current_running_case = ThreadSafeProperty('')
    current_pass_rate = ThreadSafeProperty('0%')

    @staticmethod
    def change_dos_to_unix(filepath):
        """ 文件制符变成unix"""
        with open(filepath, "rb+") as file_handle:
            all_text = file_handle.read()
            all_change_text = all_text.replace(b"\r\n", b"\n")
            file_handle.seek(0, 0)
            file_handle.truncate()
            file_handle.write(all_change_text)

    @staticmethod
    def get_filename_from_dir(dir_path, include):
        """
        获取给定文件夹下的文件名(带后缀)列表
        Args:
            dir_path: 文件夹路径
            include: 包含词汇
        Returns:
            文件名列表
        """
        files = os.listdir(dir_path)
        filenames = []
        if files:
            for filename in files:
                if include in filename:
                    filenames.append(filename)
            return filenames
        else:
            return False

    @staticmethod
    def print_run_info(run_info):
        def process_list(lst, connector_str='', step_length=2):
            if lst:
                result = []
                for i in range(0, len(lst), step_length):
                    group = lst[i: i + step_length]
                    result.append(connector_str.join(group))
                return '\n'.join(result)
            return ''

        def reformat_string(_str, step_length=20):
            if _str and isinstance(_str, str):
                # 先省略
                format_str = _str if len(_str) <= 40 else _str[:38] + '..'
                # 再换行切割
                return '\n'.join([format_str[i:i + step_length] for i in range(0, len(format_str), step_length)])
            return ''

        print(f'***{run_info.tc_name}: {run_info.tc_title}***')
        table = PrettyTable()
        table.field_names = ['行号', '结果', '描述', '行动', '预期', '实际']
        for tc_step in run_info.tc_steps:
            table.add_row([
                tc_step.row_number,
                tc_step.step_ret,
                reformat_string(tc_step.heading),
                process_list(tc_step.actions),
                process_list(tc_step.pass_condition, step_length=1),
                process_list(tc_step.evaluation_condition, step_length=1, connector_str=';')
            ])
        # 这里可以以后再做成更好看的，现在已经挺好看的了
        # table_html_str = table.get_html_string()
        # logger.info(table_html_str)
        table_str = table.get_string()
        print(table_str)
        print('\n')

    @classmethod
    def feishu_notice(cls, *args, **kwargs):
        if cls.total_num != 0:
            cls.pass_rate = '{:.1%}'.format(cls.pass_num / cls.total_num)
        else:
            cls.pass_rate = '0%'
        cls.template_bg = 'green' if cls.pass_num == cls.total_num else 'red'
        cost_seconds = int(time.time() - time.mktime(time.strptime(cls.start_time, "%Y-%m-%d %H:%M:%S")))
        m, s = divmod(cost_seconds, 60)
        h, m = divmod(m, 60)
        cls.cost_time = f'{h}小时{m}分{s}秒'
        # html_url = 'https:'  # TODO 做个local文件服务器

        template_variable = {
            # 'template_bg':cls.template_bg,  # 这里不知道模板变量为什么不生效
            'title': cls.title,
            'case_name': cls.case_name,
            'report_dir': cls.report_dir,
            # 'html_url': cls.html_url,  # 后面做本地文件服务器
            'vin': cls.vin,
            'start_time': cls.start_time,
            'cost_time': cls.cost_time,
            'result': cls.result_str,
            'total_num': str(cls.total_num),
            'pass_num': str(cls.pass_num),
            'fail_num': str(cls.fail_num),
            'error_num': str(cls.error_num),
            'pass_rate': cls.pass_rate,
            'test_detail': cls.test_detail,
        }
        payload = json.dumps({
            "receive_id": cls.chat_id,
            "msg_type": "interactive",
            "content": json.dumps(
                {
                    'type': 'template',
                    'data': {
                        'template_id': cls.card_temp_id,
                        'template_variable': template_variable,
                    }
                }
            )
        })
        headers = {
            'Content-Type': 'application/json'
        }
        params = {
            'receive_id_type': 'chat_id'
        }
        response = requests.request("POST", cls.notice_url, params=params, headers=headers, data=payload, timeout=5)
        logger.info(response.text)


class CaseParser:
    """
    解析一个excel(一个模块)下的所有用例
    """
    def __init__(self, tc_filepath):
        self.tc_filepath = tc_filepath
        self.wb = load_workbook(filename=tc_filepath)

    def parse_run_sheet(self, tc_sheet_name) -> list:
        """
        解析要执行的sheet
        :param tc_sheet_name:
        :return:
        """
        tc_sheet_name = str(tc_sheet_name).replace(' ', '').strip()
        sheet_numbers = []
        sheet_ranges = tc_sheet_name.split(";")

        for sheet_range in sheet_ranges:
            # 如果range中包含"-"符号，则解析出开始和结束的数字
            if "-" in sheet_range:
                start, end = map(int, sheet_range.split("-"))
                # 将开始和结束的数字之间的所有数字加入sheet_numbers
                for num in range(start, end + 1):
                    sheet_numbers.append(num)
            # 如果range为空，则跳过
            elif sheet_range == "":
                continue
            else:
                sheet_numbers.append(int(sheet_range))
        # 将数字转换为sheet名称，例如"3"转换为"Sheet3"
        sheet_names = ["TC" + str(num) for num in sheet_numbers]
        return sheet_names

    def get_all_testcase(self) -> dict:
        """
        获取所有测试用例的数据，以字典形式返回。
        返回结构示例如下：
            {
                <tc_sheet_name>: {
                    "case_name": str,             # 用例名称
                    "test_steps": List[TestStep]  # 测试步骤对象构成的列表
                },
                ...
            }
        """
        option_sheet = self.wb['Options']
        if option_sheet is None:
            logger.error('not found Options sheet')
            return {}
        tc_case = option_sheet['D9'].value
        if tc_case is None or tc_case == '':
            logger.error('not found testcase tc')

        all_tc = {}
        tc_case_sheets = self.parse_run_sheet(tc_case)
        for tc_name in tc_case_sheets:
            tc_info = {
                'case_name': '',
                'test_steps': []
            }
            try:
                sheet = self.wb[tc_name]
            except Exception as e:
                logger.error(f'{e} {self.tc_filepath}: {tc_name} sheet not exists')
                continue
            test_steps = []
            for row_number, row in enumerate(sheet.iter_rows(), start=1):
                if row[1].value == 'Test case' and not tc_info['case_name']:
                    tc_info['case_name'] = str(row[2].value).strip()
                if row[1].value == 'Test step':
                    actions = str(row[4].value).split('\n')
                    pass_condition = str(row[6].value).split('\n')
                    test_step = TestStep(
                        pre_condition=row[3].value,
                        actions=actions,
                        hold_condition=row[7].value,
                        wait_condition=row[5].value,
                        pass_condition=pass_condition,
                        row_number=row_number,
                        heading=row[2].value
                    )
                    test_steps.append(test_step)
            tc_info['test_steps'] = test_steps
            all_tc[tc_name] = tc_info
        return all_tc


class CaseTester:
    non_dds_prefix = (
            'SIL_',  'Sw_HandWakeup', 'sql3_', 'A2M_', 'M2A_', 'vss',
            'ssh_', 'http', 'bsp_', 'eid_fid_', 'db_', 'doip_', 'var_',
            'cal_',
        )  # 非dds消息格式

    def __init__(
            self,
            sub_topics: Optional[list] = None,
            pub_topics: Optional[list] = None,
            sdc_connector: Optional[SDCConnector] = None,
            dds_connector: Optional[DDSConnector] = None,
            ssh_connector: Optional[SSHConnector] = None,
            ssh_async_connector: Optional[SSHAsyncConnector] = None,
            doip_simulator: Optional[DoIPMonitorThread] = None,
            db_connector: Optional[DBConnector] = None,
            cloud_connector: Optional[CloudConnector] = None,
            doipclient: Optional[DoIPClient] = None,
            xcp_connector: Optional[XCPConnector] = None,
    ):
        self.sub_topics = sub_topics
        self.pub_topics = pub_topics
        self.sdc_connector = sdc_connector
        self.dds_connector = dds_connector
        self.ssh_connector = ssh_connector
        self.ssh_async_connector = ssh_async_connector
        self.doip_simulator = doip_simulator  # 新增UDS仿真
        self.db_connector = db_connector
        self.cloud_connector = cloud_connector
        self.doipclient = doipclient
        self.xcp_connector = xcp_connector
        self.callback = None

    def set_callback(self, callback):
        """
        设置pyqt主窗口回调，用户发送信号给槽函数
        Args:
            callback:

        Returns:

        """
        self.callback = callback

    @staticmethod
    def convert_signal_value(value):
        """
        数据类型转换, 这里二进制字符串、十六进制字符串、整型、浮点型，均转为浮点型； 其余支持json类型反序列化成字符串、字典、列表
        :param value:
        :return:
        """

        def replacer(match):
            var_name = match.group(1)
            return str(Variable.get_vars().get(var_name, f'{{{{{var_name}}}}}'))  # 变量不存在时保留原标记

        if value.startswith("0b") or value.startswith("0B"):
            try:
                return float(int(value[2:], 2))
            except ValueError:
                logger.error('转换二进制失败')
        elif value.startswith("0x") or value.startswith("0X"):
            try:
                return float(int(value[2:], 16))
            except ValueError:
                logger.error('转换十六进制失败')
        elif bool(re.search(r'\{\{\s*[\w]+\s*\}\}', str(value))):  # 含{{}}结构
            try:
                pattern = re.compile(r'\{\{\s*([\w]+)\s*\}\}')  # 精准匹配变量名（字母/数字/下划线）
                return pattern.sub(replacer, str(value))
            except:
                logger.error(f'解析变量值异常: {value}')
        else:
            try:
                return float(value)
            except ValueError:
                try:
                    return json.loads(value.replace('\n', ''))
                except JSONDecodeError:
                    return value

    def send_single_msg(self, signal: Variable, async_mode=True):
        """
        发送单一信号，通常为非dds信号组
        Args:
            signal: 信号，Variable实例
            async_mode: 异步模式，True为异步发送，False为同步发送，自动化通常指定False，手动调试避免ui阻塞，通常为True

        Returns:

        """

        # 用例默认信号 起填充作用 无实际意义
        if signal.name in ['Sw_HandWakeup', 'SIL_Client_CnnctSt', 'SIL_Client_Cnnct']:
            Variable(signal.name).Value = 1
        # 车辆模式ECU仿真状态
        elif signal.name.startswith('SIL_VMS_'):
            logger.info(f'设置车辆模式ECU仿真: {signal.name} = {signal.Value}')
            VehicleModeDiagnostic.set_state(signal.name[8:], signal.Value)
        # 读sqlite数据库
        elif signal.name.startswith('sql3_switch'):
            _ = signal.name.split('_')
            self.ssh_connector.get_db_signals(db_name=_[-2], table_type=_[-1])
        # 操作sqlite数据库
        elif signal.name.startswith('sql3_write_'):
            _ = signal.name.split('_')
            set_signal_name = '_'.join(_[4:])
            self.ssh_connector.set_db_signal(
                signal_name=set_signal_name,
                signal_value=signal.Value,
                db_name=_[2],
                table_type=_[3]
            )
        # 用例中要更改现有的全局变量 不作发送
        elif signal.name.startswith('A2M_'):
            logger.info(f'修改全局A2M_信号: {signal.name} = {signal.Value}')
            Variable(signal.name).Value = signal.Value
        # 发M2A TCP信号
        elif signal.name.startswith('M2A_') or signal.name.startswith('A2A_'):
            self.sdc_connector.tcp_send(signal)
        # 执行ssh终端命令
        elif signal.name.startswith('ssh_exec_cmd'):
            logger.info(f'执行ssh命令: {signal.name} = {signal.Value}')
            output = self.ssh_connector.execute_cmd(signal.Value, console=True)
            Variable('ssh_exec_output').Value = output
        # 执行ssh交互式命令，这里要改成多线程函数，否则会卡住
        elif signal.name.startswith('ssh_interactive_cmd'):
            logger.info(f'执行ssh交互式命令: {signal.name} = {signal.Value}')
            if signal.Value == 0:
                self.ssh_async_connector.interact_event.set()
            else:
                self.ssh_async_connector.execute_interact_cmd(signal.Value)
        # 执行ssh上传
        elif signal.name.startswith('ssh_sftp_put'):
            logger.info(f'执行ssh上传: {signal.name} = {signal.Value}')
            self.ssh_connector.sftp_put(
                signal.Value['local'],
                signal.Value['remote']
            )
        # 执行ssh下载
        elif signal.name.startswith('ssh_sftp_get'):
            logger.info(f'执行ssh下载: {signal.name} = {signal.Value}')
            self.ssh_connector.sftp_get(
                signal.Value['remote'],
                signal.Value['local']
            )
        # 无实际意义
        elif signal.name.startswith('ssh_exec_output'):
            pass
        # 发vss消息，这里要改成多线程函数，否则会卡住
        elif signal.name.startswith('vssSet_'):
            signal_name = signal.name[7:]
            self.ssh_connector.set_vss_signal(signal_name, signal.Value)
        # 读vss消息
        elif signal.name.startswith('vss_get'):
            self.ssh_connector.get_vss_signal(signal.Value)
        # http请求信号
        elif signal.name.startswith('httpReq_'):
            self.cloud_connector.fetch(signal.name, signal.Value, async_mode=async_mode)
        # EID FID故障信号
        elif signal.name.startswith('eid_fid_'):
            self.ssh_connector.fault_inject(signal)
        # 数据库操作
        elif signal.name.startswith('db_'):
            self.db_connector.execute(signal)
        # doip诊断请求
        elif signal.name.startswith('doip_'):
            self.doipclient.send_msg(signal)
        # 标定
        elif signal.name.startswith('cal_'):
            self.xcp_connector.send_msg(signal)
        # 临时变量存储
        elif signal.name.startswith('var_'):
            logger.info(f'变量赋值: {signal.name} = {signal.Value}')
        # 发DDS信号
        else:
            if not Variable.check_existence(signal.name):
                raise Exception(f'SignalFormatError: {signal.name}')
            else:
                self.dds_connector.dds_send(signal)

    @classmethod
    def resolve_existing_signal_name(cls, expression):
        # 使用正则表达式提取包括::和.的变量名
        pattern = r'\b[a-zA-Z_][\w\.:]*\b'
        matches = re.findall(pattern, expression)
        logger.debug(matches)
        if not matches:
            raise Exception("未找到任何合法变量名")

        for i in matches:
            # 检查提取的变量名是否在全局变量列表中
            if Variable.check_existence(i):
                variable_name = i
                return variable_name
        raise Exception(f'未找到有效的信号名')

    @classmethod
    def check_signal_name(cls, signal_name_str):
        """
        这里只检查dds（强制）和sdc信号（根据配置）
        Args:
            signal_name_str: 信号名
        Returns:
        """
        # 首先判断是不是dds信号
        if not any(signal_name_str.startswith(prefix) for prefix in cls.non_dds_prefix):
            return Variable.check_existence(signal_name_str)
        # 其次判断是不是sdc信号
        elif any(signal_name_str.startswith(prefix) for prefix in ['M2A_', 'A2A_', 'A2M_']):
            if 'check_signal_name' in env.configs and env.configs.get('check_signal_name') is False:
                return True
            else:
                return Variable.check_existence(signal_name_str)
        else:
            return True

    def run_test_case(self, tc_name, test_steps, tc_title=''):
        tc_ret = True
        test_info = TestInfo(tc_name, test_steps, tc_ret, tc_title=tc_title)
        # logger.info(f'执行测试用例 {tc_name}')
        for _i, step in enumerate(test_steps):
            if step.pre_condition:
                logger.info(f'Precondition {step.pre_condition}')
            step_evaluation = []
            test_steps[_i].test_time = datetime.now().strftime('%H:%M:%S.%f')[:-3]

            # action 输入
            action_pass = True
            # actions如果是多个，默认为一组dds信号/一组车模式仿真信号/TCP信号
            if len(step.actions) > 1:
                signals = []
                topic_name = ''
                for action in step.actions:
                    action = action.lstrip()
                    if not action.startswith(('ssh_', 'db_')):  # ssh命令可能会有空格或其他符号
                        action = action.replace(' ', '').replace('==', '=')
                    action = action.rstrip(';')
                    if action:  # 有可能是空字符串
                        _m = action.split('=')
                        sigal_name_str = _m[0]
                        # 根据信号名查找全局信号，判断是否存在
                        has_such_signal = self.check_signal_name(sigal_name_str)
                        if not has_such_signal:
                            fail_reason = f'SignalNotFound: {sigal_name_str}'
                            if fail_reason not in step_evaluation:
                                step_evaluation.append(fail_reason)
                            logger.error(f'Actions not found signal {sigal_name_str}')
                            action_pass = False
                            break

                        else:
                            try:
                                signal = Variable(sigal_name_str)
                                # 添加DDS信号组
                                if not any(signal.name.startswith(prefix) for prefix in self.non_dds_prefix):
                                    tpn = self.dds_connector.signal_map[signal.name]
                                    if topic_name and topic_name != tpn:
                                        raise Exception(f'SignalsTopicNotSame')
                                    else:
                                        topic_name = tpn
                                    signal_value = CaseTester.convert_signal_value(_m[1])
                                    if signal_value is None:
                                        logger.error(f'Signal {signal.name} value {signal_value}convert error')
                                        raise Exception(f'Signal {signal.name} value {signal_value}convert error')
                                    signal.Value = signal_value
                                    signals.append(signal)

                                # 发送车模式仿真信号
                                elif signal.name.startswith('SIL_VMS_'):
                                    signal.Value = float(_m[1])  # 字符串类型转成整型
                                    self.send_single_msg(signal, async_mode=False)

                                # 发送M2A A2M、以及其他信号
                                else:
                                    try:
                                        signal = Variable(sigal_name_str)
                                        signal_value = CaseTester.convert_signal_value('='.join(_m[1:]).strip())  # 复杂的值中可能有 "="
                                        if signal_value is None:
                                            logger.error(f'{signal.name}={signal_value} value convert error')
                                            raise Exception(f'{signal.name}={signal_value} value convert error')
                                        signal.Value = signal_value
                                        self.send_single_msg(signal, async_mode=False)
                                    except Exception as e:
                                        if str(e) not in step_evaluation:  # 防止原因重复
                                            step_evaluation.append(str(e))
                                        logger.error("Actions error: " + str(e))
                                        action_pass = False

                            except Exception as e:
                                logger.error(traceback.format_exc())
                                if str(e) not in step_evaluation:  # 防止原因重复
                                    step_evaluation.append(str(e))
                                logger.error("Actions error: " + str(e))
                                action_pass = False
                                break
                # 发送一组dds消息
                if topic_name and signals:
                    self.dds_connector.dds_multi_send(
                        topic_name=topic_name,
                        signals=signals
                    )

            elif len(step.actions) == 1 and step.actions[0] and str(step.actions[0]) != 'None':
                # 发送单信号
                action = str(step.actions[0]).lstrip()
                if not action.startswith(('ssh_', 'db_')):  # ssh命令可能会有空格或其他符号
                    action = action.replace(' ', '').replace('==', '=')
                action = action.rstrip(';')
                if action:
                    # 单条消息发送处理逻辑
                    msg = action.strip().split('=')
                    sigal_name_str = msg[0].strip()
                    has_such_signal = self.check_signal_name(sigal_name_str)
                    if not has_such_signal:
                        fail_reason = f'SignalNotFound: {sigal_name_str}'
                        if fail_reason not in step_evaluation:
                            step_evaluation.append(fail_reason)
                        logger.error(f'Actions SignalNotFound {sigal_name_str}')
                        action_pass = False

                    else:
                        try:
                            signal = Variable(sigal_name_str)
                            signal_value = CaseTester.convert_signal_value('='.join(msg[1:]).strip())  # 复杂的值中可能有 "="
                            if signal_value is None:
                                logger.error(f'{signal.name}={signal_value} value convert error')
                                raise Exception(f'{signal.name}={signal_value} value convert error')
                            signal.Value = signal_value
                            self.send_single_msg(signal, async_mode=False)
                        except Exception as e:
                            if str(e) not in step_evaluation:  # 防止原因重复
                                step_evaluation.append(str(e))
                            logger.error("Actions error: " + str(e))
                            action_pass = False

            if not action_pass:
                tc_ret = False
                if step_evaluation:
                    test_steps[_i].evaluation_condition = step_evaluation
                test_steps[_i].step_ret = False
                continue  # action异常直接进行下一个测试步骤

            # wait condition 等待
            wait = step.wait_condition
            if wait is not None and wait != '':
                try:
                    wait = str(wait).replace(' ', '')
                    time.sleep(float(wait))
                except Exception as e:
                    logger.error(f'{e}. Wait condition format error: {wait}')

            # pass condition 校验
            step_ret = None
            for pass_con in step.pass_condition:
                if pass_con is not None and pass_con != '' and pass_con != 'None':
                    pass_con = str(pass_con).replace('&&', '')
                    try:
                        variable_name = self.resolve_existing_signal_name(pass_con)

                        real_value = Variable(variable_name).Value
                        step_evaluation.append(f'{variable_name}=={real_value}')

                        # "MSG_BatACChrgEngy==nan" 这种eval会报错
                        # 这里强调一下 任何值与 NaN相比都会返回False，包括 NaN == NaN
                        # 判断类型是否为 nan，用 math.isnan()
                        if 'nan' in pass_con:
                            pass_con = pass_con.replace('nan', '"nan"')
                            # NaN == NaN返回False 这里都转换为字符进行比较
                            if not isinstance(real_value, str) and not isinstance(real_value, list) and math.isnan(real_value):
                                real_value = str(real_value)
                        # EEA2.0 rti中接收的数值可能是字符串 需要转成float类型进行比较
                        if isinstance(real_value, str) and real_value != 'nan':
                            try:
                                real_value = float(real_value)
                            except ValueError:
                                try:
                                    real_value = json.loads(real_value)
                                except ValueError:
                                    pass
                            pass
                        if isinstance(real_value, list):
                            # 如果值是列表类型且元素里包括 nan 则需要转成字符串进行比较
                            temp_value = []
                            for i in real_value:
                                if isinstance(i, float) and math.isnan(i):
                                    temp_value.append('nan')
                                else:
                                    temp_value.append(i)
                            real_value = temp_value
                        # 通过替换无效字符创建一个有效的Python变量名
                        safe_variable_name = variable_name.replace('::', '__').replace('.', '_')
                        # 所有变量都在这个作用域里体现，无法在当前环境中找到，只能在该字典中找到
                        local_namespace = {'nan': math.nan}
                        exec(f"{safe_variable_name} = {repr(real_value)}", {}, local_namespace)
                        # 现在可以安全地评估表达式了
                        result = eval(pass_con.replace(variable_name, safe_variable_name), {}, local_namespace)
                        logger.info(f'通过条件: {pass_con}, 实际值: {real_value}, 结果: {result}')
                        # if array_index is None:
                        #     logger.info(f'>>> {variable_name}: 实际值：{real_value}, 预期值：{pass_value}, 结果：{result}')
                        # else:
                        #     logger.info(f'>>> {variable_name}[{array_index}]: 实际值：{real_value[array_index]}, 预期值：{pass_value}, 结果：{result}')
                        if step_ret is None:
                            step_ret = result
                        else:
                            step_ret = result & step_ret
                        test_steps[_i].step_ret = step_ret

                    except Exception as e:
                        logger.error(traceback.format_exc())
                        evaluation_condition = str(e)
                        logger.error("Condition error: {}".format(evaluation_condition))
                        if evaluation_condition not in step_evaluation:
                            step_evaluation.append(evaluation_condition)
                        step_ret = False
                        test_steps[_i].step_ret = step_ret

                    # 一条测试步骤不过则整个用例不过
                    if not step_ret:
                        tc_ret = False
            test_steps[_i].evaluation_condition = step_evaluation
        test_info.tc_steps = test_steps
        test_info.tc_ret = tc_ret
        return test_info


class TestPrecondition:
    """
    初始化connector的连接，包括拉齐、部署，需要抛出异常给QSplash捕获失败弹窗
    """

    def __init__(self, tester: CaseTester, callback=None):
        self.tester = tester
        self.callback = callback  # pyqt progress信号

    def verify_topic_correctness(self):
        # 先判断用户给的topic列表在不在矩阵中
        if self.callback:
            self.callback.emit('检查所给topic名称是否符合矩阵要求 ...')
        logger.info('检查所给topic名称是否符合矩阵要求')
        not_matched_sub_topics = []
        not_matched_pub_topics = []
        for i in self.tester.sub_topics:
            if i not in self.tester.dds_connector.reader_topic_names:
                not_matched_sub_topics.append(i)
        for i in self.tester.pub_topics:
            if i not in self.tester.dds_connector.writer_topic_names:
                not_matched_pub_topics.append(i)
        if not_matched_sub_topics or not_matched_pub_topics:
            raise Exception(
                f'topic名称未在矩阵当中, 请重新配置yaml文件, 检查并给出正确的Topic\n'
                f'---------------------------------------------------------\n'
                f'订阅Topic: {", ".join(not_matched_sub_topics)}\n\n'
                f'发布Topic: {", ".join(not_matched_pub_topics)}'
            )
        else:
            logger.success('配置文件topic名称检查通过')

    def start_sdc_connector(self):
        if env.deploy_sil and self.tester.sdc_connector and not self.tester.sdc_connector.client_socket:
            if self.callback:
                self.callback.emit('sdc接收器 连接sil仿真节点并接收tcp消息 ...')
            self.tester.sdc_connector.connect_server()
            self.tester.sdc_connector.start()
            logger.info('>>> sdc接收器 启动tcp接收线程 ...')
            time.sleep(1)

    def start_dds_connector(self):
        """
        dds connector订阅与发布池启动
        Bug to fix: 先发布后启动，部分Topic如 ACSetStatus中的信号无法正常接受
        这一块以后再分析原因
        """
        if self.callback:
            self.callback.emit('dds订阅器 启动订阅线程池 ...')
        logger.info('>>> dds订阅器 启动订阅线程池 ...')
        for sub_tp_name in self.tester.sub_topics:
            if self.callback:
                self.callback.emit(f'订阅topic: {sub_tp_name}')
            self.tester.dds_connector.create_subscriber(sub_tp_name)
            time.sleep(0.1)
        time.sleep(1)

        if self.callback:
            self.callback.emit('dds发布器 启动发布线程池 ...')
        logger.info('>>> dds发布器 启动发布线程池 ...')
        for pub_tp_name in self.tester.pub_topics:
            if self.callback:
                self.callback.emit(f'创建publisher: {pub_tp_name}')
            self.tester.dds_connector.create_publisher(pub_tp_name)
            time.sleep(0.1)
        time.sleep(1)

    def start_ssh_connector(self):
        if self.tester.ssh_connector:
            self.tester.ssh_connector.initialize(callback=self.callback)

    def start_ssh_async_connector(self):
        if self.tester.ssh_async_connector and not self.tester.ssh_async_connector.is_alive():
            self.tester.ssh_async_connector.daemon = True
            self.tester.ssh_async_connector.start()

    def start_doip_simulator(self):
        if not self.tester.doip_simulator.is_alive():
            if self.callback:
                self.callback.emit(f'DoIP仿真线程初始化 ...')
            logger.info('>>> DoIP仿真线程初始化 ...')
            self.tester.doip_simulator.daemon = True
            self.tester.doip_simulator.start()

    def start_doipclient(self):
        if self.tester.doipclient and not self.tester.doipclient.socket_handler:
            if self.callback:
                self.callback.emit(f'DoIPClient线程初始化 ...')
            logger.info('>>> DoIPClient线程初始化 ...')
            self.tester.doipclient.connect()

    def start_xcp_connector(self):
        if self.tester.xcp_connector and not self.tester.xcp_connector.master:
            if self.callback:
                self.callback.emit(f'XCP连接器初始化 ...')
            logger.info('>>> XCP连接器初始化 ...')
            self.tester.xcp_connector.start()
            self.tester.xcp_connector.load_a2l()

    def run(self):
        self.verify_topic_correctness()
        self.start_ssh_connector()
        self.start_ssh_async_connector()
        self.start_sdc_connector()
        self.start_dds_connector()
        self.start_doipclient()
        self.start_doip_simulator()
        self.start_xcp_connector()


class TestPostCondition:
    def __init__(self, tester: CaseTester):
        self.tester = tester

    def sdc_connector_leave(self):
        if self.tester.sdc_connector and hasattr(self.tester.sdc_connector, 'client_socket'):
            self.tester.sdc_connector.stop()

    def dds_connector_leave(self):
        self.tester.dds_connector.release_connector()

    def ssh_connector_leave(self):
        if self.tester.ssh_connector:
            self.tester.ssh_connector.uninitialize()

    def ssh_async_connector_leave(self):
        if self.tester.ssh_async_connector.is_alive():
            self.tester.ssh_async_connector.close_event.set()

    def doip_simulator_leave(self):
        if self.tester.doip_simulator.is_alive():
            self.tester.doip_simulator.stop()

    def doipclient_leave(self):
        if self.tester.doipclient:
            self.tester.doipclient.close()

    def db_connector_leave(self):
        if self.tester.db_connector:
            self.tester.db_connector.close_all()

    def xcp_connector_leave(self):
        if self.tester.xcp_connector and self.tester.xcp_connector.master:
            self.tester.xcp_connector.stop()

    def run(self):
        """
        这些pyqt主线程负责的子线程，必须得先一个个安全退出，主动释放
        否则关闭窗口时报错误码
        Returns:

        """
        self.doip_simulator_leave()
        self.doipclient_leave()
        self.ssh_connector_leave()
        self.ssh_async_connector_leave()
        self.db_connector_leave()
        self.sdc_connector_leave()
        self.xcp_connector_leave()
        time.sleep(1)
        self.dds_connector_leave()

